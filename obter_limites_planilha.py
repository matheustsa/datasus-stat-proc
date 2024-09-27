import openpyxl

def obter_limites_planilha(ws):
    last_row = ws.max_row
    last_col = ws.max_column
    return last_row, last_col

def planilha_esta_vazia(ws):
    # Verifica se todas as células da primeira linha até a última têm conteúdo vazio
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value is not None:
                return False
    return True

# Exemplo de uso
wb = openpyxl.load_workbook('planilhas.xlsx')
for ws in wb.worksheets:
    linhas, colunas = obter_limites_planilha(ws)
    print(f"Tamanho da planilha '{ws.title}': {linhas} x {colunas} (linhas x colunas)")
    
    if planilha_esta_vazia(ws):
        print(f'A planilha "{ws.title}" está vazia.')
    elif linhas < 3:
        print(f'A planilha "{ws.title}" tem menos que 3 linhas.')
