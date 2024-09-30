import openpyxl

def abrir_arquivo(arquivo):
    """
    Abre um arquivo XLSX com o nome especificado.

    Args:
        arquivo (str): Nome do arquivo a ser aberto.

    Returns:
        workbook: Objetos de trabalho da planilha.
        
    Example:
        ### Abre arquivo "planilhas.xlsx"
        ```planilha = abrir_arquivo('planilhas.xlsx')
    """

    # Tenta criar um objeto de trabalho para o arquivo
    try:
        workbook = openpyxl.load_workbook(arquivo)
        
        # Caso tenha sucesso, retorna o objeto de trabalho
        return workbook
    
    except PermissionError:
        print(f"Erro: Não é possível abrir o arquivo '{arquivo}'. Ele pode estar em uso ou não existir.")
    
    except FileNotFoundError:
        print(f"Erro: O arquivo '{arquivo}' não foi encontrado. Certifique-se que ele está na mesma pasta do script ou especifique um caminho completo.")

def obter_limites_planilha(planilha):
    """
    Retorna a posição da última linha e coluna da planilha.

    Args:
        ws (worksheet): Planilha a ser consumida.
    
    Returns:
        ultima_linha: Última linha com dados na tabela.
        ultima_coluna: Última coluna com dados na tabela.
        
    Example:
        ### Pegando limites da tabela planilhas.xlsx
        ```wb = openpyxl.load_workbook('planilhas.xlsx')
        for ws in wb.worksheets:
            linhas, colunas = obter_limites_planilha(ws)
            print(f"Tamanho da planilha '{ws.title}': {linhas} x {colunas} (linhas x colunas)")
    """
    ultima_linha = planilha.max_row
    ultima_coluna = planilha.max_column
    return ultima_linha, ultima_coluna

def buscar_celula_por_texto(planilha, texto):
    """
    Busca pelo texto especificado numa célula da planilha.

    Args:
        texto (str): Texto a ser buscado.
    
    Returns:
        cell: Primeira célula da planilha onde o texto foi encontrado.
        
    Example:
        ### Buscando texto "PERIODO"
        ```wb = openpyxl.load_workbook('planilhas.xlsx')
        texto = 'PERIODO'
        for ws in wb.worksheets:
            celula_encontrada = buscar_celula_por_texto(ws, texto) 
            
            if celula_encontrada:
                print(f'Celula encontrada: {celula_encontrada.coordinate}, valor: {celula_encontrada.value}')
            else:
                print(f"Célula com o texto {texto} não foi encontrada.")
    """
    
    texto_lower = texto.lower()
    for row in planilha.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and texto_lower in cell.value.lower():
                return cell
    return None

def adicionar_celula(planilha, linha, coluna, valor=None, formatacao=None):
    """
    Adiciona uma célula a uma planilha do Excel.

    Args:
        planilha (Worksheet): A planilha onde a célula será adicionada.
        linha (int): O número da linha onde a célula será inserida.
        coluna (int): O número da coluna onde a célula será inserida.
        valor (any, optional): O valor a ser inserido na célula. Padrão é None.
        formatacao (str, optional): O formato numérico da célula (ex: '0.00%', '#,##0.00'). Padrão é None.

    Returns:
        Cell: A célula criada ou modificada.

    Example:
        ### Criando uma nova planilha e adicionando uma célula formatada
        ```wb = Workbook()
        ws = wb.active
        adicionar_celula(ws, linha=1, coluna=1, valor=0.1234, formatacao='0.00%')
        wb.save('exemplo.xlsx')
    """
    celula = planilha.cell(row=linha, column=coluna, value=valor)
    if formatacao:
        celula.number_format = formatacao
    return celula

def formata_celula(planilha, linha, coluna, formatacao):
    celula = planilha.cell(row=linha, column=coluna)
    celula.number_format = formatacao
    return celula

def copiar_planilha_com_formatacao(planilha_origem, nova_planilha):
    """
    Copia células e suas formatações de uma planilha para outra.
    """
    for row in planilha_origem.iter_rows():
        for cell in row:
            nova_celula = nova_planilha.cell(row=cell.row, column=cell.column, value=cell.value)
            
            # Copiando a formatação da célula
            if cell.has_style:
                # nova_celula.font = cell.font
                # nova_celula.border = cell.border
                # nova_celula.fill = cell.fill
                nova_celula.number_format = cell.number_format
                # nova_celula.protection = cell.protection
                nova_celula.alignment = cell.alignment

def aplicar_formatacao_coluna(planilha, coluna, formatacao):
    """
    Aplica uma formatação numérica a todas as células de uma coluna especificada.

    Args:
        planilha (Worksheet): A planilha onde a formatação será aplicada.
        coluna (int): O número da coluna onde a formatação será aplicada.
        formatacao (str): A formatação numérica a ser aplicada (ex: '0.00%', '#,##0.00').

    Example:
        ### Aplicando formatação à coluna 3
        ```wb = openpyxl.load_workbook('planilhas.xlsx')
        ws = wb.active
        aplicar_formatacao_coluna(ws, coluna=3, formatacao='0.00%')
        wb.save('planilhas_formatadas.xlsx')
    """
    ultima_linha = planilha.max_row
    for linha in range(1, ultima_linha + 1):
        celula = planilha.cell(row=linha, column=coluna)
        celula.number_format = formatacao

def remover_coluna(planilha, coluna):
    """
    Remove uma coluna especificada da planilha.

    Args:
        planilha (Worksheet): A planilha de onde a coluna será removida.
        coluna (int): O número da coluna a ser removida.

    Example:
        ### Removendo a coluna 3
        ```wb = openpyxl.load_workbook('planilhas.xlsx')
        ws = wb.active
        remover_coluna(ws, coluna=3)
        wb.save('planilhas_sem_coluna.xlsx')
    """
    planilha.delete_cols(coluna)

def renomear_planilha(planilha, titulo):
    """
    Renomeia uma planilha existente com o novo título especificado.

    Args:
        planilha (Worksheet): A planilha que será renomeada.
        titulo (str): O novo título a ser atribuído à planilha.

    Example:
        ### Renomeando uma planilha chamada "Planilha1" para "Resumo"
        ```wb = openpyxl.load_workbook('planilhas.xlsx')
        ws = wb['Planilha1']
        renomear_planilha(ws, 'Resumo')
        wb.save('planilhas_renomeadas.xlsx')
    """
    planilha.title = titulo
