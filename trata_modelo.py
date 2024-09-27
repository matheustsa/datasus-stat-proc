import openpyxl

def carregar_planilha(nome_arquivo):
    return openpyxl.load_workbook(nome_arquivo)

def buscar_celula_por_texto(ws, texto_parcial):
    texto_parcial_lower = texto_parcial.lower()
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and texto_parcial_lower in cell.value.lower():
                return cell
    return None

def obter_limites_planilha(ws):
    last_row = ws.max_row
    last_col = ws.max_column
    return last_row, last_col

def adiciona_colunas_dados_obtidos_e_ausentes(ws, linha_cabecalhos, last_col):
    dados_obtidos = ws.cell(row=linha_cabecalhos, column=last_col + 1)
    dados_obtidos.value = 'DADOS_OBTIDOS (%)'
    
    dados_ausentes = ws.cell(row=linha_cabecalhos, column=last_col + 2)
    dados_ausentes.value = 'DADOS_AUSENTES (%)'
    
    return dados_obtidos, dados_ausentes

def calcula_porcentagens_dados_obtidos_e_ausentes(ws, table_data_start_cell, table_data_last_cell, dados_obtidos, dados_ausentes):
    for row in range(table_data_start_cell.row, table_data_last_cell.row + 1):
        intervalo_linha = f'{table_data_start_cell.column_letter}{row}:{table_data_last_cell.column_letter}{row}'
        
        porcent_obtidos = ws.cell(row=row, column=dados_obtidos.col_idx)
        porcent_obtidos.value = f'=COUNT({intervalo_linha}) / COUNTA({intervalo_linha})'
        porcent_obtidos.number_format = '0.00%'
        
        porcent_ausentes = ws.cell(row=row, column=dados_ausentes.col_idx)
        porcent_ausentes.value = f'=1-{porcent_obtidos.column_letter}{row}'
        porcent_ausentes.number_format = '0.00%'

def calcula_porcentagens_gerais(ws, table_data_start_cell, table_data_last_cell, last_row):
    ws.cell(row=last_row + 1, column=1).value = '-------------'
    
    porcent_completa = ws.cell(row=last_row + 1, column=1)
    porcent_completa.value = 'PORCENT_COMPLETA (%)'
    intervalo_dados = f'{table_data_start_cell.coordinate}:{table_data_last_cell.coordinate}'
    
    valor_porcent_completa = ws.cell(row=porcent_completa.row, column=2)
    valor_porcent_completa.value = f'=COUNT({intervalo_dados}) / COUNTA({intervalo_dados})'
    valor_porcent_completa.number_format = '0.00%'
    
    porcent_ausente = ws.cell(row=last_row + 2, column=1)
    porcent_ausente.value = 'PORCENT_AUSENTE (%)'
    
    valor_porcent_ausente = ws.cell(row=porcent_ausente.row, column=2)
    valor_porcent_ausente.value = f'=1-{valor_porcent_completa.column_letter}{valor_porcent_completa.row}'
    valor_porcent_ausente.number_format = '0.00%'

def processar_planilha(ws):
    last_row, last_col = obter_limites_planilha(ws)
    
    celula_periodo = buscar_celula_por_texto(ws, 'Periodo')
    linha_cabecalhos = celula_periodo.row + 1
    
    table_data_start_cell = ws.cell(row=linha_cabecalhos + 1, column=2)
    table_data_last_row = last_row - 1
    table_data_last_cell = ws.cell(row=table_data_last_row, column=last_col)
    
    dados_obtidos, dados_ausentes = adiciona_colunas_dados_obtidos_e_ausentes(ws, linha_cabecalhos, last_col)
    calcula_porcentagens_dados_obtidos_e_ausentes(ws, table_data_start_cell, table_data_last_cell, dados_obtidos, dados_ausentes)
    calcula_porcentagens_gerais(ws, table_data_start_cell, table_data_last_cell, last_row)

def processar_planilhas(nome_arquivo_entrada, nome_arquivo_saida):
    wb = carregar_planilha(nome_arquivo_entrada)
    quant_planilhas = len(wb.worksheets)
    quant_planilhas_processadas = 0
    planilhas_nao_processadas = []

    for ws in wb.worksheets:
        if ws.max_row > 3:
            processar_planilha(ws)
            quant_planilhas_processadas += 1
        else:
            planilhas_nao_processadas.append(ws.title)

    wb.save(nome_arquivo_saida)
    print(f'Planilhas processadas: {quant_planilhas_processadas}/{quant_planilhas}')
    
    # Criar arquivo de texto com planilhas não processadas
    if planilhas_nao_processadas:
        with open('planilhas_nao_processadas.txt', 'w', encoding='utf-8') as file:
            file.write('Planilhas não processadas:\n')
            for planilha in planilhas_nao_processadas:
                file.write(f'{planilha}\n')

        print(f'Lista de planilhas não processadas salva em "planilhas_nao_processadas.txt"')
    else:
        print('Todas as planilhas foram processadas com sucesso!')


# Executa o processamento
processar_planilhas('planilhas.xlsx', 'planilhas_tratadas.xlsx')
