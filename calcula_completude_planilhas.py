import openpyxl
from openpyxl_helper import *

# Carregar o arquivo existente e criar um novo arquivo
planilhas_tratadas_xlsx = openpyxl.load_workbook('planilhas_tratadas.xlsx')
novo_arquivo_xlsx = openpyxl.Workbook()
nova_planilha = novo_arquivo_xlsx.active
nova_planilha.title = "Completude das planilhas"

# Adicionar o cabeçalho
nova_planilha.append(['Completude das planilhas'])
nova_planilha.append(['Nome da planilha', 'Porcentagem completa', 'Porcentagem ausente'])

# Processar cada planilha
for planilha in planilhas_tratadas_xlsx.worksheets:
    cel_percent_completa = buscar_celula_por_texto(planilha, 'PORCENT_COMPLETA (%)')
    valor_percent_completa = cel_percent_completa.offset(0, 1).value
    valor_percent_ausente = cel_percent_completa.offset(1, 1).value
    nova_planilha.append([planilha['A1'].value, valor_percent_completa, valor_percent_ausente])

# Adicionar separador
nova_planilha.append(['-------------'])

# Coletar os valores das porcentagens completas e ausentes
intervalo_completa = [
    nova_planilha.cell(linha, 2).value for linha in range(3, nova_planilha.max_row)
]
intervalo_ausente = [
    nova_planilha.cell(linha, 3).value for linha in range(3, nova_planilha.max_row)
]

# Calcular médias de completude e ausência
valor_percent_completa = sum(intervalo_completa) / len(intervalo_completa)
valor_percent_ausente = sum(intervalo_ausente) / len(intervalo_ausente)

# Adicionar os resultados calculados à planilha
nova_planilha.append(['PORCENT_COMPLETA (%)', valor_percent_completa])
nova_planilha.append(['PORCENT_AUSENTE (%)', valor_percent_ausente])



# Definir o formato percentual
for linha in range(3, nova_planilha.max_row + 1):
    nova_planilha.cell(linha, 2).number_format = '0.00%'
    nova_planilha.cell(linha, 3).number_format = '0.00%'

# Salvar o arquivo final
novo_arquivo_xlsx.save('completude_planilhas_tratadas.xlsx')
