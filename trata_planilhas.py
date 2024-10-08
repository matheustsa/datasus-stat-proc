from openpyxl import Workbook
import argparse
from openpyxl_helper import *

class PlanilhaProcessor:
    """
    Classe responsável por processar planilhas Excel, adicionando colunas com
    cálculos de porcentagens de dados obtidos e ausentes.

    Attributes:
        nome_arquivo_entrada (str): Nome do arquivo Excel de entrada.
        nome_arquivo_saida (str): Nome do arquivo Excel de saída.
        lista_planilhas_nao_processadas (list): Lista de planilhas que não foram processadas.
        arquivo_xlsx (Workbook): Objeto que representa o arquivo Excel aberto.
    
    Examples:
        ### Exemplo de execução do script sobre arquivo "planilhas.xlsx"
        ```$> python trata_planilhas.py planilhas.xlsx planilhas_tratadas.xlsx
    """

    def __init__(self, nome_arquivo_entrada, nome_arquivo_saida):
        """
        Inicializa o Processador de Planilhas com o nome do arquivo de entrada e saída.

        Args:
            nome_arquivo_entrada (str): Nome do arquivo Excel de entrada.
            nome_arquivo_saida (str): Nome do arquivo Excel de saída.
        """

        self.cabecalho_periodo = 'Período'
        self.cabecalho_dados_obtidos = 'DADOS_OBTIDOS (%)'
        self.cabecalho_dados_ausentes = 'DADOS_AUSENTES (%)'
        self.cabecalho_procent_completa = 'PORCENT_COMPLETA (%)'
        self.cabecalho_procent_ausente = 'PORCENT_AUSENTE (%)'

        self.number_format_porcentagem = '0.00%'

        self.nome_arquivo_entrada = nome_arquivo_entrada
        self.nome_arquivo_saida = nome_arquivo_saida
        self.lista_planilhas_nao_processadas = []
        self.arquivo_xlsx = self.abrir_arquivo()
        self.novo_arquivo_xlsx = Workbook()
        self.novo_arquivo_xlsx.remove(self.novo_arquivo_xlsx.active)

    def abrir_arquivo(self):
        """
        Abre o arquivo Excel para processamento.

        Returns:
            Workbook: O arquivo Excel aberto.
        """
        return abrir_arquivo(self.nome_arquivo_entrada)

    def processar_planilha(self, planilha):
        """
        Processa uma única planilha adicionando colunas de porcentagem de dados obtidos
        e dados ausentes, além de calcular as porcentagens gerais.

        Args:
            planilha (Worksheet): Planilha a ser processada.
        """
        print(f'Processando a planilha: {planilha.title}...')

        ultima_linha_planilha, ultima_coluna_planilha = obter_limites_planilha(planilha)

        # Verifica se a planilha é válida
        if ultima_linha_planilha < 3:
            print('Planilha inválida.')
            self.lista_planilhas_nao_processadas.append(planilha.title)
            return

        celula_periodo = buscar_celula_por_texto(planilha, self.cabecalho_periodo)
        if not celula_periodo:
            print('Planilha inválida.')
            self.lista_planilhas_nao_processadas.append(planilha.title)
            return

        # Definindo o intervalo de dados e as colunas adicionais
        linha_cabecalhos = celula_periodo.row + 1
        inicio_dados_tabela = planilha.cell(row=linha_cabecalhos + 1, column=2)
        fim_dados_tabela = planilha.cell(row=ultima_linha_planilha - 1, column=ultima_coluna_planilha)

        cel_dados_obtidos, cel_dados_ausentes = self.adicionar_colunas_dados_obtidos_ausentes(planilha, linha_cabecalhos, ultima_coluna_planilha)
        self.calcular_porcentagens_individuais(planilha, inicio_dados_tabela, fim_dados_tabela, cel_dados_obtidos, cel_dados_ausentes)
        
        self.calcular_porcentagens_gerais(planilha, inicio_dados_tabela, fim_dados_tabela, ultima_linha_planilha)

        # cria nova planilha no novo_arquivo_xlsx e copia os dados da planilha
        nova_planilha = self.novo_arquivo_xlsx.create_sheet(planilha.title)
        for row in planilha.iter_rows(values_only=True):
            nova_planilha.append(row)
        self.formata_planilha(nova_planilha)

    def adicionar_colunas_dados_obtidos_ausentes(self, planilha, linha_cabecalhos, ultima_coluna):
        """
        Adiciona colunas de 'DADOS_OBTIDOS' e 'DADOS_AUSENTES' à planilha.

        Args:
            planilha (Worksheet): A planilha onde as colunas serão adicionadas.
            linha_cabecalhos (int): Linha onde os cabeçalhos estão localizados.
            ultima_coluna (int): A última coluna com dados na planilha.

        Returns:
            tuple: Colunas onde 'DADOS_OBTIDOS' e 'DADOS_AUSENTES' foram inseridos.
        """
        cel_dados_obtidos = adicionar_celula(planilha, linha=linha_cabecalhos, coluna=ultima_coluna + 1, valor=self.cabecalho_dados_obtidos)
        cel_dados_ausentes = adicionar_celula(planilha, linha=linha_cabecalhos, coluna=ultima_coluna + 2, valor=self.cabecalho_dados_ausentes)
        return cel_dados_obtidos, cel_dados_ausentes

    def calcular_porcentagens_individuais(self, planilha, inicio_dados_tabela, fim_dados_tabela, cel_dados_obtidos, cel_dados_ausentes):
        """
        Calcula as porcentagens de dados obtidos e ausentes para cada linha da tabela.

        Args:
            planilha (Worksheet): A planilha onde os dados serão calculados.
            inicio_dados_tabela (Cell): Célula que marca o início dos dados.
            fim_dados_tabela (Cell): Célula que marca o fim dos dados.
            col_dados_obtidos (Cell): Coluna onde a porcentagem de dados obtidos será inserida.
            col_dados_ausentes (Cell): Coluna onde a porcentagem de dados ausentes será inserida.
        """

        for linha in range (inicio_dados_tabela.row, fim_dados_tabela.row + 1):
            intervalo = [planilha.cell(linha, coluna).value for coluna in range(inicio_dados_tabela.col_idx, cel_dados_obtidos.col_idx)]

            valores_validos = [valor for valor in intervalo if isinstance(valor, (int, float))]
            
            media = len(valores_validos) / len(intervalo)
            diferenca = 1 - media
            
            adicionar_celula(planilha, linha, cel_dados_obtidos.col_idx, media, self.number_format_porcentagem)
            adicionar_celula(planilha, linha, cel_dados_ausentes.col_idx, diferenca, self.number_format_porcentagem)

    def calcular_porcentagens_gerais(self, planilha, inicio_dados_tabela, fim_dados_tabela, ultima_linha_planilha):
        """
        Calcula as porcentagens gerais de dados obtidos e ausentes para a tabela inteira.

        Args:
            planilha (Worksheet): A planilha onde os cálculos serão feitos.
            inicio_dados_tabela (Cell): Célula que marca o início dos dados.
            fim_dados_tabela (Cell): Célula que marca o fim dos dados.
            ultima_linha_planilha (int): A última linha da planilha.
        """
        # Adiciona separadores e cabeçalhos
        adicionar_celula(planilha, linha=ultima_linha_planilha + 1, coluna=1, valor='-------------')
        adicionar_celula(planilha, linha=ultima_linha_planilha + 2, coluna=1, valor=self.cabecalho_procent_completa)
        adicionar_celula(planilha, linha=ultima_linha_planilha + 3, coluna=1, valor=self.cabecalho_procent_ausente)

        # Obter os valores da planilha em um único passo usando list comprehension
        intervalo = [
            planilha.cell(linha, coluna).value
            for linha in range(inicio_dados_tabela.row, fim_dados_tabela.row + 1)
            for coluna in range(inicio_dados_tabela.col_idx, fim_dados_tabela.col_idx + 1)
        ]

        # Filtrar apenas valores numéricos (int ou float)
        valores_validos = [valor for valor in intervalo if isinstance(valor, (int, float))]

        # Calcular a média de porcentagem e diferença
        quant_valores_intervalo = len(intervalo)
        media = len(valores_validos) / quant_valores_intervalo if quant_valores_intervalo > 0 else 0
        diferenca = 1 - media

        # Adicionar os resultados na planilha
        adicionar_celula(planilha, ultima_linha_planilha + 2, 2, media)
        adicionar_celula(planilha, ultima_linha_planilha + 3, 2, diferenca)


    def salvar_planilhas_nao_processadas(self):
        """
        Salva as planilhas que não foram processadas em um arquivo de texto.
        """
        if self.lista_planilhas_nao_processadas:
            with open('planilhas_nao_processadas.txt', 'w', encoding='utf-8') as file:
                file.write('Planilhas não processadas:\n')
                for planilha in self.lista_planilhas_nao_processadas:
                    file.write(f'{planilha}\n')
            
            print(f'Lista de planilhas não processadas ({self.lista_planilhas_nao_processadas.__len__()}) salva em "planilhas_nao_processadas.txt"')
        else:
            print('Todas as planilhas foram processadas com sucesso!')      

    def formata_planilha(self, planilha):
        celula_dados_obtidos = buscar_celula_por_texto(planilha, self.cabecalho_dados_obtidos)
        celula_dados_ausentes = celula_dados_obtidos.offset(0,1)
        aplicar_formatacao_coluna(planilha, coluna=celula_dados_obtidos.column, formatacao=self.number_format_porcentagem)
        aplicar_formatacao_coluna(planilha, coluna=celula_dados_ausentes.column, formatacao=self.number_format_porcentagem)
        
        celula_porcent_completa = buscar_celula_por_texto(planilha, self.cabecalho_procent_completa)
        celula_porcent_ausente = celula_porcent_completa.offset(1,0)
        aplicar_formatacao_linha(planilha, celula_porcent_completa.row, self.number_format_porcentagem)
        aplicar_formatacao_linha(planilha, celula_porcent_ausente.row, self.number_format_porcentagem)

    def processar(self):
        """
        Processa planilhas do Excel.
        """
        
        for _, planilha in enumerate(self.arquivo_xlsx.worksheets):
            self.processar_planilha(planilha)

        print('Salvando planilha gerada...')
        self.novo_arquivo_xlsx.save(self.nome_arquivo_saida)
        print('Salvando lista de planilhas não processadas...')
        self.salvar_planilhas_nao_processadas()
        print('Processamento concluído.')


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Processar planilhas Excel adicionando porcentagens de dados.')
    parser.add_argument('nome_arquivo_entrada', help='Nome do arquivo Excel de entrada')
    parser.add_argument('nome_arquivo_saida', help='Nome do arquivo Excel de saída')
    
    args = parser.parse_args()
    
    processor = PlanilhaProcessor(args.nome_arquivo_entrada, args.nome_arquivo_saida)
    processor.processar()
